import os
import csv
import io
import datetime
from sqlalchemy.orm import Session as DBSession
from fpdf import FPDF
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from models.finance import Income, Expense, ExpenseCategory
from models.budget import Budget, SavingsGoal, Debt, SinkingFund
from models.household import Household
from utils.helpers import format_currency

# Create a reports folder within workspace if it doesn't exist
REPORTS_DIR = "reports_cache"
os.makedirs(REPORTS_DIR, exist_ok=True)

class PDFReport(FPDF):
    def header(self):
        # Header style
        self.set_font('Helvetica', 'B', 16)
        self.set_text_color(0, 199, 255)  # Cyan theme color
        self.cell(0, 10, 'SMARTBUDGET AI - FINANCIAL SUMMARY REPORT', border=0, ln=1, align='L')
        self.set_draw_color(0, 199, 255)
        self.line(10, 20, 200, 20)
        self.ln(10)
        
    def footer(self):
        # Position at 1.5 cm from bottom
        self.set_y(-15)
        self.set_font('Helvetica', 'I', 8)
        self.set_text_color(128, 128, 128)
        self.cell(0, 10, f'Page {self.page_no()} | Generated on {datetime.date.today().strftime("%d %b %Y")} | SmartBudget AI', 0, 0, 'C')

def generate_pdf_report(db: DBSession, household_id: int) -> str:
    """
    Generates a beautifully designed financial summary PDF report.
    Returns the file path of the generated report.
    """
    household = db.query(Household).filter(Household.id == household_id).first()
    currency = household.currency if household else "FJD"
    
    # 1. Gather stats
    incomes = db.query(Income).filter(Income.household_id == household_id).all()
    expenses = db.query(Expense).filter(Expense.household_id == household_id).all()
    debts = db.query(Debt).filter(Debt.household_id == household_id).all()
    goals = db.query(SavingsGoal).filter(SavingsGoal.household_id == household_id).all()
    
    total_income = sum(i.amount for i in incomes)
    total_expense = sum(e.amount for e in expenses)
    net_savings = total_income - total_expense
    
    # Init PDF
    pdf = PDFReport()
    pdf.add_page()
    pdf.set_font("Helvetica", size=11)
    
    # Executive Summary Card
    pdf.set_fill_color(240, 248, 255) # Light blue card background
    pdf.rect(10, 25, 190, 45, 'F')
    
    pdf.set_y(28)
    pdf.set_font("Helvetica", "B", 12)
    pdf.set_text_color(26, 36, 43)
    pdf.cell(10)
    pdf.cell(0, 6, f"Household Profile: {household.name if household else 'Default'}", ln=1)
    
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(10)
    pdf.cell(0, 5, f"Budget Method: {household.budget_method.upper()} | Currency Code: {currency}", ln=1)
    
    pdf.ln(5)
    pdf.cell(10)
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(50, 5, f"Total Income: {format_currency(total_income, currency)}")
    pdf.cell(50, 5, f"Total Spent: {format_currency(total_expense, currency)}")
    pdf.cell(50, 5, f"Net Balance: {format_currency(net_savings, currency)}", ln=1)
    
    pdf.ln(12)
    
    # Category Spending Table
    pdf.set_font("Helvetica", "B", 12)
    pdf.set_text_color(0, 150, 200)
    pdf.cell(0, 8, "Spending Breakdown by Category", ln=1)
    pdf.set_draw_color(200, 200, 200)
    pdf.set_line_width(0.2)
    
    # Group expenses
    cat_spend = {}
    for exp in expenses:
        cat_name = exp.category.name if exp.category else "Other"
        cat_spend[cat_name] = cat_spend.get(cat_name, 0.0) + exp.amount
        
    # Table Header
    pdf.set_font("Helvetica", "B", 10)
    pdf.set_fill_color(230, 235, 240)
    pdf.set_text_color(50, 50, 50)
    pdf.cell(100, 8, "Category", border=1, fill=True)
    pdf.cell(90, 8, "Amount", border=1, ln=1, fill=True, align="R")
    
    pdf.set_font("Helvetica", "", 10)
    for cat, amount in sorted(cat_spend.items(), key=lambda x: x[1], reverse=True):
        pdf.cell(100, 7, cat, border=1)
        pdf.cell(90, 7, format_currency(amount, currency), border=1, ln=1, align="R")
        
    pdf.ln(8)
    
    # Debt & Goals Section
    if debts:
        pdf.set_font("Helvetica", "B", 12)
        pdf.set_text_color(0, 150, 200)
        pdf.cell(0, 8, "Outstanding Household Debts", ln=1)
        
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(80, 8, "Debt Name", border=1, fill=True)
        pdf.cell(40, 8, "Interest Rate", border=1, fill=True, align="C")
        pdf.cell(70, 8, "Current Balance", border=1, ln=1, fill=True, align="R")
        
        pdf.set_font("Helvetica", "", 10)
        for d in debts:
            pdf.cell(80, 7, d.name, border=1)
            pdf.cell(40, 7, f"{d.interest_rate}%", border=1, align="C")
            pdf.cell(70, 7, format_currency(d.current_balance, currency), border=1, ln=1, align="R")
            
        pdf.ln(8)
        
    if goals:
        pdf.set_font("Helvetica", "B", 12)
        pdf.set_text_color(0, 150, 200)
        pdf.cell(0, 8, "Savings Goals Progress", ln=1)
        
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(80, 8, "Goal", border=1, fill=True)
        pdf.cell(55, 8, "Target Amount", border=1, fill=True, align="R")
        pdf.cell(55, 8, "Current Savings", border=1, ln=1, fill=True, align="R")
        
        pdf.set_font("Helvetica", "", 10)
        for g in goals:
            pdf.cell(80, 7, g.name, border=1)
            pdf.cell(55, 7, format_currency(g.target_amount, currency), border=1, align="R")
            pdf.cell(55, 7, format_currency(g.current_amount, currency), border=1, ln=1, align="R")
            
    # Save file
    filename = f"report_summary_{household_id}_{datetime.date.today().isoformat()}.pdf"
    filepath = os.path.join(REPORTS_DIR, filename)
    pdf.output(filepath)
    return filepath

def generate_excel_report(db: DBSession, household_id: int) -> str:
    """
    Generates a structured, multi-sheet Excel spreadsheet report with budget variances.
    Returns the file path of the generated spreadsheet.
    """
    household = db.query(Household).filter(Household.id == household_id).first()
    currency = household.currency if household else "FJD"
    
    wb = openpyxl.Workbook()
    
    # 1. Summary Sheet
    ws_summary = wb.active
    ws_summary.title = "Financial Summary"
    
    # Styling variables
    font_title = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    font_section = Font(name="Calibri", size=12, bold=True)
    font_bold = Font(name="Calibri", size=11, bold=True)
    fill_header = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid") # Dark navy
    fill_row = PatternFill(start_color="F2F4F8", end_color="F2F4F8", fill_type="solid")
    
    ws_summary.merge_cells("A1:D1")
    ws_summary["A1"] = "SmartBudget AI - Household Executive Summary"
    ws_summary["A1"].font = font_title
    ws_summary["A1"].fill = fill_header
    ws_summary["A1"].alignment = Alignment(horizontal="center")
    
    ws_summary["A3"] = "Household Name:"
    ws_summary["A3"].font = font_bold
    ws_summary["B3"] = household.name if household else "Default"
    ws_summary["A4"] = "Generated At:"
    ws_summary["A4"].font = font_bold
    ws_summary["B4"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    
    # Income & Expense lists
    incomes = db.query(Income).filter(Income.household_id == household_id).all()
    expenses = db.query(Expense).filter(Expense.household_id == household_id).all()
    
    # 2. Income Sheet
    ws_income = wb.create_sheet(title="Income Register")
    ws_income.append(["Date", "Source", "Recurring", "Amount"])
    ws_income.row_dimensions[1].font = font_bold
    for inc in incomes:
        ws_income.append([inc.date, inc.source, "Yes" if inc.is_recurring else "No", inc.amount])
        
    # 3. Expense Sheet
    ws_expense = wb.create_sheet(title="Expense Register")
    ws_expense.append(["Date", "Category", "Merchant", "Amount", "Notes"])
    ws_expense.row_dimensions[1].font = font_bold
    for exp in expenses:
        ws_expense.append([
            exp.date, 
            exp.category.name if exp.category else "Other", 
            exp.merchant or "", 
            exp.amount, 
            exp.notes or ""
        ])
        
    # Write formulas to Summary sheet
    ws_summary["A7"] = "Financial Health Overview"
    ws_summary["A7"].font = font_section
    
    ws_summary["A9"] = "Total Income:"
    ws_summary["B9"] = f"='Income Register'!D{len(incomes)+2}" if incomes else 0
    # Wait, simple sum formula is safer
    ws_summary["B9"] = "=SUM('Income Register'!D2:D5000)"
    ws_summary["B9"].font = font_bold
    
    ws_summary["A10"] = "Total Expenses:"
    ws_summary["B10"] = "=SUM('Expense Register'!D2:D5000)"
    ws_summary["B10"].font = font_bold
    
    ws_summary["A11"] = "Net Savings:"
    ws_summary["B11"] = "=B9-B10"
    ws_summary["B11"].font = font_bold
    
    # Format columns
    for col in ["B9", "B10", "B11"]:
        ws_summary[col].number_format = "$#,##0.00"
        
    # Save file
    filename = f"excel_report_{household_id}_{datetime.date.today().isoformat()}.xlsx"
    filepath = os.path.join(REPORTS_DIR, filename)
    wb.save(filepath)
    return filepath

def generate_csv_export(db: DBSession, household_id: int, export_type: str = "expenses") -> str:
    """
    Generates a raw CSV dump of expenses or incomes.
    Returns the file path.
    """
    filename = f"csv_{export_type}_{household_id}_{datetime.date.today().isoformat()}.csv"
    filepath = os.path.join(REPORTS_DIR, filename)
    
    with open(filepath, mode="w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        
        if export_type == "expenses":
            writer.writerow(["ID", "Category", "Amount", "Date", "Merchant", "Notes", "Tags"])
            expenses = db.query(Expense).filter(Expense.household_id == household_id).all()
            for exp in expenses:
                writer.writerow([
                    exp.id, 
                    exp.category.name if exp.category else "Other", 
                    exp.amount, 
                    exp.date.isoformat(), 
                    exp.merchant or "", 
                    exp.notes or "", 
                    exp.tags or ""
                ])
        else: # income
            writer.writerow(["ID", "Source", "Amount", "Date", "Recurring", "Frequency", "Description"])
            incomes = db.query(Income).filter(Income.household_id == household_id).all()
            for inc in incomes:
                writer.writerow([
                    inc.id, 
                    inc.source, 
                    inc.amount, 
                    inc.date.isoformat(), 
                    "Yes" if inc.is_recurring else "No", 
                    inc.frequency or "", 
                    inc.description or ""
                ])
                
    return filepath
